# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
 @Author  : pgsheng
 @Time    : 2018/4/15 19:47
"""
import openpyxl
import os

from openpyxl.styles import Font, colors, Alignment

from public import config

"""备份一份测试数据，以防进行写入表操作时导致表出现异常，表打不开数据丢失"""


def copy_excel(file_name):
    dir_name = os.path.dirname(file_name)  # 获取文件路径
    basename = os.path.basename(file_name)  # 获取文件名
    file_name2 = os.path.join(dir_name, "backup_" + basename)  # 备份excel的文件

    wb2 = openpyxl.Workbook()  # 创建一个excel文件
    wb2.save(file_name2)
    # 根据文件名加载文件
    wb1 = openpyxl.load_workbook(file_name)
    wb2 = openpyxl.load_workbook(file_name2)
    sheets1 = wb1.sheetnames  # 获取文件的所有工作表

    sheets1 = sheets1[::-1]  # 倒序
    for sheetName in sheets1:
        ws1 = wb1[sheetName]
        ws2 = wb2.create_sheet(sheetName, 0)  # 创建一个工作表

        max_row = ws1.max_row  # 最大行数
        max_column = ws1.max_column  # 最大列数
        for m in list(range(1, max_row + 1)):
            for n in list(range(97, 97 + max_column)):  # chr(97)='a'
                n = chr(n)  # ASCII字符
                i = '%s%d' % (n, m)  # 单元格编号
                cell1 = ws1[i].value  # 获取data单元格数据
                ws2[i].value = cell1  # 赋值到test单元格

    wb2.save(file_name2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()


class WriteExcel(object):
    # 初始化,filename是文件路径,sheetName是表名
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.wb = openpyxl.load_workbook(self.filename)  # 打开文件
        self.ws = self.wb[sheet_name]  # 根据表名打开表

    # 写入excel数据,row是行数,col是列数，value是要写入的数据,color是写入字体颜色（默认黑色）
    def write(self, row, col, value, color=colors.BLACK):
        self.ws.cell(row, col).value = value  # 写入数据到指定单元格
        self.ws.cell(row, col).font = Font(color=color)
        self.ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        self.wb.save(self.filename)  # 保存文件


if __name__ == "__main__":
    # path = copy_excel(Config.project_path + r'\test_data\test_01.xlsx', 'Sheet1')
    wt = WriteExcel(config.test_data_path + 'test_01.xlsx', 'test')
    wt.write(5, 6, "HELLEOP2", colors.RED)
    wt.write(5, 7, "HELLEOP2")
    wt.wb.close()
